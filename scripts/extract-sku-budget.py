from __future__ import annotations

import json
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "SKU-бюджет" / "Магнат_Расчет бюджета на Перформ.xlsx"
OUTPUT = ROOT / "public" / "data" / "sku-budget.json"


def value(cell: Any) -> Any:
    if isinstance(cell, (datetime, date)):
        return cell.isoformat()
    return cell


def as_float(raw: Any) -> float:
    if raw is None or raw == "":
        return 0.0
    try:
        return float(raw)
    except (TypeError, ValueError):
        return 0.0


def as_int(raw: Any) -> int:
    return int(round(as_float(raw)))


def short_name(name: str, limit: int = 82) -> str:
    clean = " ".join(str(name).split())
    if len(clean) <= limit:
        return clean
    return clean[: limit - 1].rstrip() + "..."


def num_lit_values(source: Any) -> list[float]:
    num_lit = getattr(source, "numLit", None)
    if num_lit is None:
        return []
    points = sorted(num_lit.pt, key=lambda point: point.idx)
    return [as_float(point.v) for point in points]


def series_label(series: Any) -> str:
    label = getattr(series, "tx", None)
    if label is None:
        return ""
    return str(getattr(label, "v", "") or "")


def action_from_delta(delta: float, current: float, status: str) -> str:
    if "слабая" in status.lower() and current <= 0:
        return "Нет активного бюджета"
    if current <= 0 and delta <= 0:
        return "Нет активного бюджета"
    if delta > max(5000, current * 0.05):
        return "Увеличить"
    if delta < -max(5000, current * 0.05):
        return "Снизить"
    return "Оставить"


def recommendation_from_action(action: str, recommended_daily: float, status: str) -> str:
    if action == "Увеличить":
        return f"Масштабировать поддержку до {recommended_daily:,.0f} RUB в день".replace(",", " ")
    if action == "Снизить":
        return f"Снизить дневной бюджет до {recommended_daily:,.0f} RUB".replace(",", " ")
    if action == "Нет активного бюджета":
        return "Не закладывать регулярную поддержку без отдельного теста"
    if "слабая" in status.lower():
        return "Оставить как тестовую позицию: связь бюджета и продаж слабая"
    return "Оставить текущий уровень и наблюдать за фактом"


def collect_key_values(
    ws: Any,
    label_col: int,
    value_col: int,
    start_row: int,
    end_row: int,
) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for row in range(start_row, end_row + 1):
        label = ws.cell(row, label_col).value
        raw_value = ws.cell(row, value_col).value
        if label:
            result[str(label)] = value(raw_value)
    return result


def main() -> None:
    workbook = load_workbook(SOURCE, data_only=True)

    summary_sheet = workbook["Сводка 3д"]
    checks_sheet = workbook["Проверки"]
    sku_sheet = workbook["SKU 3д"]
    daily_sheet = workbook["Дневные точки 3д"]
    chart_sheet = workbook["График полный"]

    summary_rows = collect_key_values(summary_sheet, 1, 3, 5, 8)
    summary_metrics = collect_key_values(summary_sheet, 6, 7, 4, 14)
    checks = collect_key_values(checks_sheet, 1, 2, 3, 20)

    headers = [sku_sheet.cell(4, col).value for col in range(1, sku_sheet.max_column + 1)]
    sku_rows: dict[str, dict[str, Any]] = {}
    for row in range(5, sku_sheet.max_row + 1):
        row_values = {
            str(headers[col - 1]): sku_sheet.cell(row, col).value
            for col in range(1, len(headers) + 1)
            if headers[col - 1]
        }
        if not row_values.get("Артикул"):
            continue

        sku = str(row_values["Артикул"])
        current_month = as_float(row_values["Текущий факт/мес"])
        recommended_month = as_float(row_values["Реко бюджет/мес при текущей частоте"])
        delta_month = as_float(row_values["Разница к текущему/мес"])
        current_day = as_float(row_values["Текущий активный бюджет/день"])
        recommended_day = as_float(row_values["Реко бюджет/день"])
        status = str(row_values["Статус точки"])
        action = action_from_delta(delta_month, current_month, status)

        sku_rows[sku] = {
            "sku": sku,
            "name": str(row_values["Товар"]),
            "shortName": short_name(str(row_values["Товар"])),
            "category": str(row_values["Категория"]),
            "activeDays": as_int(row_values["Активных дней"]),
            "activeDaysPerMonth": as_float(row_values["Активных дней/мес"]),
            "skuStatMonthly": as_float(row_values["Бюджет sku_stat/мес"]),
            "generateMonthly": as_float(row_values["Бюджет generate/мес"]),
            "currentMonthly": current_month,
            "currentDaily": current_day,
            "saturationDaily": as_float(row_values["Расчетная точка насыщения/день"]),
            "recommendedDaily": recommended_day,
            "recommendedMonthly": recommended_month,
            "alwaysOnMonthly": as_float(row_values["Реко бюджет/мес если ежедневно"]),
            "deltaMonthly": delta_month,
            "drr": as_float(row_values["DRR 3д, %"]),
            "cr": as_float(row_values["CR 3д, %"]),
            "ctr": as_float(row_values["CTR 3д, %"]),
            "salesUnits3d": as_int(row_values["Продажи 3д, шт"]),
            "r2": as_float(row_values["R² модели"]),
            "status": status,
            "recommendation": recommendation_from_action(action, recommended_day, status),
            "action": action,
        }

    daily_headers = [
        daily_sheet.cell(4, col).value for col in range(1, daily_sheet.max_column + 1)
    ]
    daily_points: dict[str, list[dict[str, Any]]] = defaultdict(list)
    weighted_prices: dict[str, dict[str, float]] = defaultdict(
        lambda: {"rub": 0.0, "units": 0.0}
    )

    for row in range(5, daily_sheet.max_row + 1):
        row_values = {
            str(daily_headers[col - 1]): daily_sheet.cell(row, col).value
            for col in range(1, len(daily_headers) + 1)
            if daily_headers[col - 1]
        }
        if not row_values.get("Артикул"):
            continue

        sku = str(row_values["Артикул"])
        units = as_float(row_values["Факт продаж 3д, шт"])
        rub = as_float(row_values["Sales RUB 3д"])
        weighted_prices[sku]["rub"] += rub
        weighted_prices[sku]["units"] += units
        daily_points[sku].append(
            {
                "date": value(row_values["Дата"]),
                "budget": round(as_float(row_values["Бюджет всего"]), 2),
                "skuStatBudget": round(as_float(row_values["Бюджет sku_stat"]), 2),
                "generateBudget": round(as_float(row_values["Бюджет generate"]), 2),
                "salesUnits": round(units, 2),
                "modelUnits": round(as_float(row_values["Модель продаж 3д, шт"]), 2),
                "salesRub": round(rub, 2),
                "orders": round(as_float(row_values["Заказы"]), 2),
                "clicks": round(as_float(row_values["Клики"]), 2),
                "impressions": round(as_float(row_values["Показы"]), 2),
            }
        )

    chart_series: dict[str, list[dict[str, float]]] = {}
    for chart in chart_sheet._charts:
        for series in chart.series:
            sku = series_label(series)
            if not sku:
                continue
            x_values = num_lit_values(series.xVal)
            y_values = num_lit_values(series.yVal)
            if x_values and y_values and len(x_values) == len(y_values):
                chart_series[sku] = [
                    {"budget": round(x, 2), "salesUnits": round(y, 2)}
                    for x, y in zip(x_values, y_values)
                ]

    for sku, sku_data in sku_rows.items():
        price_bucket = weighted_prices[sku]
        avg_price = (
            price_bucket["rub"] / price_bucket["units"]
            if price_bucket["units"]
            else 0.0
        )
        curve = chart_series.get(sku)
        curve_with_rub = (
            [
                {
                    "budget": point["budget"],
                    "salesUnits": point["salesUnits"],
                    "salesRub": round(point["salesUnits"] * avg_price, 2),
                }
                for point in curve
            ]
            if curve
            else []
        )
        sku_data["averageUnitPrice"] = round(avg_price, 2)
        sku_data["points"] = daily_points.get(sku, [])
        sku_data["curve"] = curve_with_rub

    data = {
        "id": "sku-budget",
        "title": "SKU-бюджет: оптимальный performance-бюджет Ozon",
        "description": (
            "Модель показывает, как дневной performance-бюджет связан с продажами "
            "SKU Магнат, где находится текущий уровень и какую точку бюджета "
            "разумно заложить в медиаплан."
        ),
        "source": "SKU-бюджет/Магнат_Расчет бюджета на Перформ.xlsx",
        "period": "2026-01-01 .. 2026-06-06",
        "methodology": {
            "budgetRule": "Общий performance-бюджет SKU за день = sku_stat + generate_products.",
            "salesRule": "Продажи 3д = sell-out SKU в штуках за день поддержки и следующие 2 календарных дня.",
            "model": "Sales = c + a * (1 - exp(-k * Budget)); точка насыщения = 85% от асимптоты.",
            "limitation": "Это historical response-curve, а не causal test: на продажи также влияют скидка, наличие, органика, конкуренты и позиция в выдаче.",
        },
        "overview": {
            "currentMonthly": as_float(summary_rows["Текущий факт"]),
            "recommendedMonthly": as_float(summary_rows["Оптимум"]),
            "testCeilingMonthly": as_float(summary_rows["Потолок теста"]),
            "alwaysOnMonthly": as_float(summary_rows["Если поддерживать ежедневно"]),
            "deltaMonthly": as_float(summary_metrics["Разница к текущему, RUB/мес"]),
            "skuCount": as_int(summary_metrics["SKU в расчете/модели"]),
            "dailyPoints": as_int(summary_metrics["Дневных точек"]),
            "excludedNoImpressions": as_int(
                summary_metrics["Дней с бюджетом без показов исключено"]
            ),
            "confirmedPlateau": as_int(checks.get("SKU с подтвержденным плато", 0)),
            "aboveObservedPlateau": as_int(
                checks.get("SKU с плато выше наблюдаемого диапазона", 0)
            ),
            "weakRelation": as_int(checks.get("SKU со слабой связью", 0)),
            "meanR2": as_float(checks["Средний R²"]),
            "medianR2": as_float(checks["Медианный R²"]),
        },
        "skus": list(sku_rows.values()),
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT.relative_to(ROOT)} with {len(data['skus'])} SKU")


if __name__ == "__main__":
    main()
