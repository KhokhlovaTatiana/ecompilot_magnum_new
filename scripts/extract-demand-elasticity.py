from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "Эластичность" / "Магнат эластичность цены.xlsx"
OUTPUT = ROOT / "public" / "data" / "demand-elasticity.json"


SCENARIO_LABELS = {
    "Baseline": "База",
    "Price -10%": "Цена -10%",
    "Price -20%": "Цена -20%",
    "Price +10%": "Цена +10%",
    "Price +20%": "Цена +20%",
}

READOUT_LABELS = {
    "Current level": "Текущий уровень",
    "Lower price, more volume": "Снижение цены, рост объема",
    "Lower price, strong lift": "Снижение цены, сильный рост объема",
    "Higher price, lower volume": "Повышение цены, падение объема",
    "Higher price, stronger drop": "Повышение цены, сильное падение объема",
}


def as_float(value: Any, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    return float(value)


def as_int(value: Any) -> int:
    return int(round(as_float(value)))


def format_date(value: Any) -> str:
    text = str(value)
    if " " in text:
        text = text.split(" ", 1)[0]
    return text


def short_label(value: str, limit: int = 96) -> str:
    text = " ".join(str(value).split())
    if len(text) <= limit:
        return text
    return text[: limit - 1].rstrip() + "..."


def sensitivity_label(elasticity: float) -> str:
    abs_value = abs(elasticity)
    if abs_value >= 2.5:
        return "Очень высокая чувствительность"
    if abs_value >= 1:
        return "Эластичный спрос"
    return "Неэластичный спрос"


def reliability_label(r_squared: float) -> str:
    if r_squared >= 0.4:
        return "Модель с хорошей объясняющей силой"
    if r_squared >= 0.25:
        return "Модель средней устойчивости"
    return "Модель требует осторожной интерпретации"


def read_summary(sheet: Any) -> dict[str, Any]:
    return {
        sheet.cell(row=row, column=1).value: sheet.cell(row=row, column=2).value
        for row in range(1, sheet.max_row + 1)
        if sheet.cell(row=row, column=1).value
    }


def read_scenarios(sheet: Any) -> list[dict[str, Any]]:
    headers = [
        sheet.cell(row=4, column=col).value for col in range(1, sheet.max_column + 1)
    ]
    rows: list[dict[str, Any]] = []

    for row in range(5, sheet.max_row + 1):
        scenario_name = sheet.cell(row=row, column=1).value
        if not scenario_name or str(scenario_name).startswith("Interpretation"):
            continue

        values = {
            header: sheet.cell(row=row, column=index + 1).value
            for index, header in enumerate(headers)
            if header
        }
        rows.append(
            {
                "scenario": SCENARIO_LABELS.get(str(scenario_name), str(scenario_name)),
                "priceChange": as_float(values.get("Price change")),
                "newAveragePrice": as_float(values.get("New avg price, RUB")),
                "salesMultiplier": as_float(values.get("Sales multiplier"), 1),
                "expectedSalesChange": as_float(values.get("Expected sales change")),
                "expectedAnnualUnits": as_float(values.get("Expected annual units")),
                "expectedDailyUnits": as_float(values.get("Expected daily units")),
                "businessReadout": READOUT_LABELS.get(
                    str(values.get("Business readout")), str(values.get("Business readout"))
                ),
                "dailyRevenue": as_float(values.get("Выручка")),
                "marginPerUnit": None,
                "dailyMargin": None,
            }
        )

    return rows


def read_top_skus(sheet: Any) -> list[dict[str, Any]]:
    headers = [
        sheet.cell(row=4, column=col).value for col in range(1, sheet.max_column + 1)
    ]
    rows: list[dict[str, Any]] = []

    for row in range(5, sheet.max_row + 1):
        label = sheet.cell(row=row, column=1).value
        if not label:
            continue

        values = {
            header: sheet.cell(row=row, column=index + 1).value
            for index, header in enumerate(headers)
            if header
        }
        rows.append(
            {
                "label": short_label(str(values["SKU label"])),
                "barcode": str(values["Barcode"]),
                "code": str(values["Code"]),
                "category": str(values["Category level 4"]),
                "sellOutQty": as_float(values["Sell-out qty"]),
                "revenue": as_float(values["Revenue, RUB"]),
                "qtyShare": as_float(values["Share of qty"]),
                "revenueShare": as_float(values["Share of revenue"]),
                "averagePrice": as_float(values["Avg unit price, RUB"]),
                "averageDiscount": as_float(values["Avg discount, %"]),
                "daysWithSales": as_int(values["Days with sales"]),
            }
        )

    return rows


def main() -> None:
    workbook = openpyxl.load_workbook(SOURCE, data_only=True, read_only=True)
    summary = read_summary(workbook["Summary"])
    scenarios = read_scenarios(workbook["Scenarios"])
    top_skus = read_top_skus(workbook["Top SKUs"])
    workbook.close()

    start = format_date(summary["Window start"])
    end = format_date(summary["Window end"])
    elasticity = as_float(summary["Price elasticity"])
    r_squared = as_float(summary["R-squared"])
    plus_10 = next(row for row in scenarios if row["priceChange"] == 0.1)[
        "expectedSalesChange"
    ]
    minus_10 = next(row for row in scenarios if row["priceChange"] == -0.1)[
        "expectedSalesChange"
    ]

    category = {
        "key": "icecream",
        "title": "Мороженое Магнат",
        "shortTitle": "Мороженое",
        "source": "Эластичность/Магнат эластичность цены.xlsx",
        "description": "Дневной sell-out бренда Магнат на Ozon за 12 месяцев.",
        "period": f"{start} — {end}",
        "summary": {
            "selectedSkus": len(top_skus),
            "selectedBarcodes": len({sku["barcode"] for sku in top_skus}),
            "selectedCodes": len({sku["code"] for sku in top_skus}),
            "daysWithSales": as_int(summary["Days with sell-out"]),
            "daysUsedInRegression": as_int(summary["Days used in regression"]),
            "totalUnits": as_float(summary["Total sell-out units"]),
            "totalRevenue": as_float(summary["Total revenue, RUB"]),
            "averagePrice": as_float(summary["Weighted avg unit price, RUB"]),
            "averageDiscount": as_float(summary["Weighted avg discount, %"]),
            "elasticity": elasticity,
            "rSquared": r_squared,
            "correlation": 0,
            "impactPlus10": plus_10,
            "impactMinus10": minus_10,
            "interpretation": (
                "Эластичность ниже -1: при повышении цены объем продаж заметно "
                "снижается, а промо-снижение цены может дать прирост штук."
            ),
        },
        "assessment": {
            "label": sensitivity_label(elasticity),
            "reliability": reliability_label(r_squared),
            "text": (
                "Модель показывает эластичный спрос: сценарии цены стоит оценивать "
                "вместе с маржинальностью и промо-календарем, особенно для SKU с "
                "крупной долей выручки."
            ),
        },
        "summaryRows": [
            {"label": "Период", "value": f"{start} — {end}"},
            {"label": "SKU в выборке", "value": len(top_skus), "format": "number"},
            {
                "label": "Дней в регрессии",
                "value": as_int(summary["Days used in regression"]),
                "format": "number",
            },
            {
                "label": "Продажи, шт",
                "value": as_float(summary["Total sell-out units"]),
                "format": "number",
            },
            {
                "label": "Выручка",
                "value": as_float(summary["Total revenue, RUB"]),
                "format": "currency",
            },
            {
                "label": "Средняя цена",
                "value": as_float(summary["Weighted avg unit price, RUB"]),
                "format": "currency",
            },
            {
                "label": "Средняя скидка",
                "value": as_float(summary["Weighted avg discount, %"]),
                "format": "percentPoint",
            },
            {"label": "Эластичность цены", "value": elasticity, "format": "decimal"},
            {"label": "R²", "value": r_squared, "format": "decimal"},
            {"label": "Эффект +10% цены", "value": plus_10, "format": "percent"},
            {"label": "Эффект -10% цены", "value": minus_10, "format": "percent"},
        ],
        "scenarioRows": scenarios,
        "method": (
            "Daily sell-out from sell_out_ozon; elasticity is estimated as "
            "ln(units) on ln(unit price)."
        ),
    }

    data = {
        "id": "demand-elasticity",
        "title": "Эластичность спроса Магнат",
        "description": (
            "Сценарная модель показывает, как изменение средней цены связано "
            "с ожидаемым объемом продаж и дневной выручкой."
        ),
        "source": "Эластичность/Магнат эластичность цены.xlsx",
        "period": category["period"],
        "overall": {
            "elasticity": elasticity,
            "rSquared": r_squared,
            "impactPlus10": plus_10,
            "impactMinus10": minus_10,
            "assessment": sensitivity_label(elasticity),
            "basis": "регрессия ln(units) на ln(unit price)",
            "text": (
                "При +10% к цене модель ожидает снижение продаж примерно на "
                f"{abs(plus_10) * 100:.1f}%, а при -10% к цене - рост объема "
                f"примерно на {minus_10 * 100:.1f}%."
            ),
        },
        "categories": [category],
        "topSkus": top_skus,
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT.relative_to(ROOT)} with {len(top_skus)} top SKU rows")


if __name__ == "__main__":
    main()
